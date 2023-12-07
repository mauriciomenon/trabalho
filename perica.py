import openpyxl
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl.utils import get_column_letter
from datetime import datetime

# Definir os estilos utilizados no cabeçalho e nas células
header_font = Font(bold=True)
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def adjust_column_width(sheet):
    for col in sheet.columns:
        max_length = 0
        column = col[0].column  # Get the column name
        for cell in col:
            try:  # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = max_length + 2
        sheet.column_dimensions[get_column_letter(column)].width = adjusted_width


# Função para adicionar o cabeçalho ao relatório
def create_header(sheet, headers):
    sheet.append(headers)
    for col in range(1, len(headers) + 1):
        cell = sheet.cell(row=1, column=col)
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")
    adjust_column_width(sheet)


# Função para adicionar uma linha de dados no relatório
def add_data_row(sheet, data_row):
    sheet.append(data_row)
    row = sheet.max_row
    for col in range(1, len(data_row) + 1):
        cell = sheet.cell(row=row, column=col)
        cell.border = thin_border
    adjust_column_width(sheet)


# Função principal para criar o relatório
def create_report(filename):
    wb = openpyxl.Workbook()
    sheet = wb.active
    headers = [
        "Dia(s)",
        "Natureza da Atividade",
        "Área de Risco",
        "Local",
        "Descrição da Atividade de Risco",
        "Visto do Gerente ou Supervisor",
    ]
    create_header(sheet, headers)

    # Você adicionaria os dados aqui chamando add_data_row(sheet, data_row) para cada linha de dados

    # Salvando o arquivo
    wb.save(filename)


# Nome do arquivo incluindo o mês e ano correntes
current_month_year = datetime.now().strftime("%Y_%m")
filename = f"Relatorio_Atividades_Risco_{current_month_year}.xlsx"

create_report(filename)
